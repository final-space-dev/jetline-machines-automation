#!/usr/bin/env python3
"""
Export comparison results to Excel with multiple sheets
"""

import mysql.connector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Tuple
from collections import defaultdict
import os
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

# Database credentials
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

if not DB_USER or not DB_PASSWORD:
    raise ValueError("DB_USER and DB_PASSWORD must be set in .env file")

# SSH Tunnel Mode: Check environment variable
USE_SSH_TUNNEL = os.getenv("USE_SSH_TUNNEL", "false").lower() == "true"
SSH_TUNNEL_PORT = int(os.getenv("SSH_TUNNEL_PORT", "3307"))

if USE_SSH_TUNNEL:
    print(f"✓ SSH Tunnel mode enabled (localhost:{SSH_TUNNEL_PORT})\n")

# Helper function to get connection host/port
def get_connection(host, port=3306):
    """Return connection details - localhost if SSH tunnel mode, otherwise original"""
    if USE_SSH_TUNNEL:
        # In SSH tunnel mode, ALL connections go through the tunnel to localhost
        # The jump server (172.20.246.163) can reach all database servers
        return ("localhost", SSH_TUNNEL_PORT)
    return (host, port)

# Companies list (excluding braamfonteinbms2)
COMPANIES = [
    {"schema": "corporateprintbms2", "name": "Jetline Corporate", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "burlingtonbms2", "name": "Burlington", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "typoprintingbms2", "name": "Typo", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "anglobms2", "name": "Anglo", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "formattsolutionsbms2", "name": "Formatt", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "landkbms2", "name": "L and K", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "marinsbms2", "name": "Marins", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "25amcpsbms2", "name": "Corporate Print 25AM", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "masterskillbms2", "name": "Masterskill", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "nscbms2", "name": "Fixtrade", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "pocketmediabms2", "name": "Pocket Media", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "raptorbms2", "name": "Raptor", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "systemprintbms2", "name": "SystemPrint", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "firstlabelsbms2", "name": "First Labels", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "welkombms2", "name": "Welkom", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "witbankbms2", "name": "Witbank", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "gardensbms2", "name": "Gardens", "host": get_connection("wizardzgardens.jetlinestores.co.za")[0], "port": get_connection("wizardzgardens.jetlinestores.co.za")[1]},
    {"schema": "printoutsolutionsbms2", "name": "PrintOut", "host": get_connection("172.20.251.127")[0], "port": get_connection("172.20.251.127")[1]},
    {"schema": "waterfrontbms2", "name": "Waterfront", "host": get_connection("wizardzwaterfront.jetlinestores.co.za")[0], "port": get_connection("wizardzwaterfront.jetlinestores.co.za")[1]},
    {"schema": "centurycitybms2", "name": "Century City", "host": get_connection("centurycity.jetlinestores.co.za")[0], "port": get_connection("centurycity.jetlinestores.co.za")[1]},
    {"schema": "albertonbms2", "name": "Alberton", "host": get_connection("alberton.jetlinestores.co.za")[0], "port": get_connection("alberton.jetlinestores.co.za")[1]},
    {"schema": "bedfordviewbms2", "name": "Bedfordview", "host": get_connection("bedfordview.jetlinestores.co.za")[0], "port": get_connection("bedfordview.jetlinestores.co.za")[1]},
    {"schema": "blackheathbms2", "name": "Blackheath", "host": get_connection("blackheath.jetlinestores.co.za")[0], "port": get_connection("blackheath.jetlinestores.co.za")[1]},
    {"schema": "boksburgbms2", "name": "Boksburg", "host": get_connection("boksburg.jetlinestores.co.za")[0], "port": get_connection("boksburg.jetlinestores.co.za")[1]},
    {"schema": "benonibms2", "name": "Benoni", "host": get_connection("benoni.jetlinestores.co.za")[0], "port": get_connection("benoni.jetlinestores.co.za")[1]},
    {"schema": "bryanstonbms2", "name": "Bryanston", "host": get_connection("bryanston.jetlinestores.co.za")[0], "port": get_connection("bryanston.jetlinestores.co.za")[1]},
    {"schema": "durbanbms2", "name": "Durban", "host": get_connection("durban.jetlinestores.co.za")[0], "port": get_connection("durban.jetlinestores.co.za")[1]},
    {"schema": "foxstreetbms2", "name": "Foxstreet", "host": get_connection("foxstreet.jetlinestores.co.za")[0], "port": get_connection("foxstreet.jetlinestores.co.za")[1]},
    {"schema": "parktowncorporateprintbms2", "name": "Hillcrestcps", "host": get_connection("parktown.jetlinestores.co.za")[0], "port": get_connection("parktown.jetlinestores.co.za")[1]},
    {"schema": "hillcrestbms2", "name": "Hillcrest", "host": get_connection("hillcrest.jetlinestores.co.za")[0], "port": get_connection("hillcrest.jetlinestores.co.za")[1]},
    {"schema": "kyalamibms2", "name": "Kyalami", "host": get_connection("kyalami.jetlinestores.co.za")[0], "port": get_connection("kyalami.jetlinestores.co.za")[1]},
    {"schema": "melrosebms2", "name": "Melrose", "host": get_connection("melrose.jetlinestores.co.za")[0], "port": get_connection("melrose.jetlinestores.co.za")[1]},
    {"schema": "menlynbms2", "name": "Menlyn", "host": get_connection("menlyn.jetlinestores.co.za")[0], "port": get_connection("menlyn.jetlinestores.co.za")[1]},
    {"schema": "parktownbms2", "name": "Parktown", "host": get_connection("parktown.jetlinestores.co.za")[0], "port": get_connection("parktown.jetlinestores.co.za")[1]},
    {"schema": "pietermaritzburgbms2", "name": "Pietermaritzburg", "host": get_connection("pietermaritzburg.jetlinestores.co.za")[0], "port": get_connection("pietermaritzburg.jetlinestores.co.za")[1]},
    {"schema": "sunninghillbms2", "name": "Sunninghill", "host": get_connection("sunninghill.jetlinestores.co.za")[0], "port": get_connection("sunninghill.jetlinestores.co.za")[1]},
    {"schema": "polokwanebms2", "name": "Polokwane", "host": get_connection("polokwane.jetlinestores.co.za")[0], "port": get_connection("polokwane.jetlinestores.co.za")[1]},
    {"schema": "rivoniabms2", "name": "Rivonia", "host": get_connection("rivonia.jetlinestores.co.za")[0], "port": get_connection("rivonia.jetlinestores.co.za")[1]},
    {"schema": "rosebankbms2", "name": "Rosebank", "host": get_connection("rosebank.jetlinestores.co.za")[0], "port": get_connection("rosebank.jetlinestores.co.za")[1]},
    {"schema": "rustenburgbms2", "name": "Rustenburg", "host": get_connection("rustenburg.jetlinestores.co.za")[0], "port": get_connection("rustenburg.jetlinestores.co.za")[1]},
    {"schema": "sandownbms2", "name": "Sandown", "host": get_connection("sandown.jetlinestores.co.za")[0], "port": get_connection("sandown.jetlinestores.co.za")[1]},
    {"schema": "illovobms2", "name": "Illovo", "host": get_connection("illovo.jetlinestores.co.za")[0], "port": get_connection("illovo.jetlinestores.co.za")[1]},
    {"schema": "montanabms2", "name": "Montana", "host": get_connection("montana.jetlinestores.co.za")[0], "port": get_connection("montana.jetlinestores.co.za")[1]},
    {"schema": "brooklynbms2", "name": "Brooklyn", "host": get_connection("brooklyn.jetlinestores.co.za")[0], "port": get_connection("brooklyn.jetlinestores.co.za")[1]},
    {"schema": "potchbms2", "name": "Potchefstroom", "host": get_connection("potchefstroom.jetlinestores.co.za")[0], "port": get_connection("potchefstroom.jetlinestores.co.za")[1]},
    {"schema": "woodmeadbms2", "name": "Woodmead", "host": get_connection("woodmead.jetlinestores.co.za")[0], "port": get_connection("woodmead.jetlinestores.co.za")[1]},
    {"schema": "georgebms2", "name": "George", "host": get_connection("george.jetlinestores.co.za")[0], "port": get_connection("george.jetlinestores.co.za")[1]},
    {"schema": "modderfonteinbms2", "name": "Modderfontein", "host": get_connection("modderfontein.jetlinestores.co.za")[0], "port": get_connection("modderfontein.jetlinestores.co.za")[1]},
    {"schema": "vaalreefsbms2", "name": "Vaalreefs", "host": get_connection("vaalreefs.jetlinestores.co.za")[0], "port": get_connection("vaalreefs.jetlinestores.co.za")[1]},
    {"schema": "klerksdorpbms2", "name": "Klerksdorp", "host": get_connection("klerksdorp.jetlinestores.co.za")[0], "port": get_connection("klerksdorp.jetlinestores.co.za")[1]},
    {"schema": "greenpointbms2", "name": "Greenpoint", "host": get_connection("greenpoint.jetlinestores.co.za")[0], "port": get_connection("greenpoint.jetlinestores.co.za")[1]},
    {"schema": "randburgbms2", "name": "Randburg", "host": get_connection("randburg.jetlinestores.co.za")[0], "port": get_connection("randburg.jetlinestores.co.za")[1]},
    {"schema": "hydeparkbms2", "name": "Hydepark", "host": get_connection("hydepark.jetlinestores.co.za")[0], "port": get_connection("hydepark.jetlinestores.co.za")[1]},
    {"schema": "fourwaysbms2", "name": "Fourways", "host": get_connection("fourways.jetlinestores.co.za")[0], "port": get_connection("fourways.jetlinestores.co.za")[1]},
    {"schema": "centurionbms2", "name": "Centurion", "host": get_connection("centurion.jetlinestores.co.za")[0], "port": get_connection("centurion.jetlinestores.co.za")[1]},
    {"schema": "nelspruitbms2", "name": "Nelspruit", "host": get_connection("nelspruit.jetlinestores.co.za")[0], "port": get_connection("nelspruit.jetlinestores.co.za")[1]},
    {"schema": "witsbms2", "name": "Wits", "host": get_connection("wits.jetlinestores.co.za")[0], "port": get_connection("wits.jetlinestores.co.za")[1]},
    {"schema": "constantiabms2", "name": "Constantia", "host": get_connection("constantia.jetlinestores.co.za")[0], "port": get_connection("constantia.jetlinestores.co.za")[1]},
    {"schema": "stellenboschbms2", "name": "Stellenbosch", "host": get_connection("stellenbosch.jetlinestores.co.za")[0], "port": get_connection("stellenbosch.jetlinestores.co.za")[1]},
    {"schema": "tygervalleybms2", "name": "Tygervalley", "host": get_connection("tygervalley.jetlinestores.co.za")[0], "port": get_connection("tygervalley.jetlinestores.co.za")[1]},
    {"schema": "midrandbms2", "name": "Midrand", "host": get_connection("midrand.jetlinestores.co.za")[0], "port": get_connection("midrand.jetlinestores.co.za")[1]},
    {"schema": "mmabathobms2", "name": "Mmabatho", "host": get_connection("mmabatho.jetlinestores.co.za")[0], "port": get_connection("mmabatho.jetlinestores.co.za")[1]},
    {"schema": "ballitobms2", "name": "Ballito", "host": get_connection("ballito.jetlinestores.co.za")[0], "port": get_connection("ballito.jetlinestores.co.za")[1]}
]

QUERY = """
SELECT
    ma.serialnumber,
    ma.machine_model_name,
    ma.machinestatus,
    mrd.reading_date,
    crme.createdtime as reading_date_time,
    mrd.total as totalamt,
    mrd.a3,
    mrd.black,
    mrd.large,
    mrd.colour,
    mrd.extralarge,
    mrd.meterreadingid,
    mrd.meterreading_no
FROM bms_machines ma
LEFT JOIN vtiger_crmentity crm on crm.crmid = ma.machinesid
LEFT JOIN bms_meterreading mrd on mrd.asset = ma.machinesid
LEFT JOIN vtiger_crmentity crme on crme.crmid = mrd.meterreadingid
WHERE crm.deleted <> 1
    AND ma.serialnumber IS NOT NULL
    AND mrd.total IS NOT NULL
    AND mrd.reading_date IS NOT NULL
    AND ma.machinestatus = 1
ORDER BY ma.serialnumber, mrd.reading_date, crme.createdtime;
"""


def load_device_meters(filepath: str) -> Dict:
    """Load Device Current Meters file - pivot by serial number with meter types as columns

    Logic (per Xerox aggregation rules):
    - If Total Impressions exists and > 0 → use that (ignore all else)
    - Else use Black Impressions + Color Impressions = Total
    - Ignore Black Large, Color Large, Extra Long (they're already included in Black/Color)
    - Add 'calc_method' field to indicate how total was calculated
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Structure: {serial: {company, product, last_date, black, color, total_impressions}}
    data = {}

    # Only these meter types matter
    meter_map = {
        'Black Impressions': 'black',
        'Color Impressions': 'color',
        'Total Impressions': 'total_impressions'
    }

    for row_num in range(2, ws.max_row + 1):
        company = ws.cell(row_num, 1).value
        product = ws.cell(row_num, 2).value
        serial = str(ws.cell(row_num, 3).value).strip() if ws.cell(row_num, 3).value else None
        meter_desc = ws.cell(row_num, 4).value
        reading = ws.cell(row_num, 5).value
        last_date = ws.cell(row_num, 6).value

        if not serial or meter_desc not in meter_map:
            continue

        if serial not in data:
            data[serial] = {
                'company': company,
                'product': product,
                'last_date': last_date,
                'black': 0,
                'color': 0,
                'total_impressions': 0
            }

        field = meter_map[meter_desc]
        data[serial][field] = reading or 0

        # Update last_date to the most recent across ALL meter types
        if last_date and (not data[serial]['last_date'] or last_date > data[serial]['last_date']):
            data[serial]['last_date'] = last_date

    # Calculate final total based on rules
    for serial in data:
        d = data[serial]
        total_impressions = d.get('total_impressions', 0) or 0
        black = d.get('black', 0) or 0
        color = d.get('color', 0) or 0

        if total_impressions > 0:
            # Rule 1: Total Impressions exists - use it
            d['total'] = total_impressions
            d['calc_method'] = 'Total Impressions'
        elif black > 0 or color > 0:
            # Rule 2: Sum Black + Color
            d['total'] = black + color
            d['calc_method'] = 'Black + Color'
        else:
            # No valid meters found
            d['total'] = 0
            d['calc_method'] = 'NO DATA'

    return data


def load_excel_data(filepath: str) -> Dict:
    """Load Excel data with sub-meters - filter to specific months only"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    data = defaultdict(lambda: defaultdict(dict))

    # Only these months from 2025
    allowed_months = {'Mar-2025', 'Apr-2025', 'May-2025', 'Jun-2025',
                     'Jul-2025', 'Aug-2025', 'Sep-2025', 'Nov-2025'}

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer, model, serial, month_year, a3_mono, a4_mono, a3_color, a4_color, total = row
        if serial and month_year in allowed_months:
            data[str(serial)][month_year] = {
                'customer': customer,
                'a3_mono': a3_mono or 0,
                'a4_mono': a4_mono or 0,
                'a3_color': a3_color or 0,
                'a4_color': a4_color or 0,
                'total': total or 0
            }

    return data


def calculate_incremental_volumes(raw_data: List[Tuple]) -> Dict:
    """
    Convert cumulative to incremental with sub-meters - matches Qlik logic
    Takes last reading per date (FirstSortedValue), then calculates monthly movement
    Process ALL historical data but only return 2025 months for reporting
    """
    from collections import OrderedDict

    # Only report these months (but calculate from beginning of time)
    report_months = {'Mar-2025', 'Apr-2025', 'May-2025', 'Jun-2025',
                     'Jul-2025', 'Aug-2025', 'Sep-2025', 'Nov-2025'}

    # Step 1: Group by serial+date, take LAST reading per date (by createdtime)
    by_serial_date = defaultdict(lambda: defaultdict(list))

    # raw_data structure: (serial, model, status, date, date_time, total, a3, black, large, colour, extralarge, meterreading_id, meterreading_no, company)
    for serial, model, status, date, date_time, total, a3, black, large, colour, extralarge, meterreading_id, meterreading_no, company in raw_data:
        if serial and date and total is not None:
            date_key = date.strftime("%Y-%m-%d")
            by_serial_date[str(serial)][date_key].append((
                date_time, total, a3 or 0, black or 0, large or 0, colour or 0, extralarge or 0, model, company, meterreading_id
            ))

    # Step 2: Get latest reading per date
    daily_readings = defaultdict(dict)
    for serial, dates in by_serial_date.items():
        for date_key, readings in dates.items():
            # Sort by datetime desc, take first (latest)
            readings.sort(key=lambda x: x[0] if x[0] else datetime.min, reverse=True)
            latest = readings[0]
            daily_readings[serial][date_key] = {
                'total': latest[1],
                'a3': latest[2],
                'black': latest[3],
                'large': latest[4],
                'colour': latest[5],
                'extralarge': latest[6],
                'model': latest[7],
                'company': latest[8],
                'meterreading_id': latest[9]
            }

    # Step 3: Group by month, take last day's reading of month
    result = defaultdict(lambda: defaultdict(dict))

    for serial, dates in daily_readings.items():
        sorted_dates = sorted(dates.items(), key=lambda x: x[0])
        monthly_readings = OrderedDict()

        for date_str, reading in sorted_dates:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            month_year = date_obj.strftime("%b-%Y")
            # Take latest day in month
            monthly_readings[month_year] = reading

        # Step 4: Calculate incremental movement from ALL history, but only store 2025 months
        # Get list of months to iterate with next month lookup
        month_list = list(monthly_readings.keys())

        for i, month_year in enumerate(month_list):
            reading = monthly_readings[month_year]

            # Get previous month's reading for baseline
            if i > 0:
                prev_reading = monthly_readings[month_list[i-1]]
                prev_total = prev_reading['total']
                prev_a3 = prev_reading['a3']
                prev_black = prev_reading['black']
                prev_large = prev_reading['large']
                prev_colour = prev_reading['colour']
                prev_extralarge = prev_reading['extralarge']
            else:
                prev_total = 0
                prev_a3 = 0
                prev_black = 0
                prev_large = 0
                prev_colour = 0
                prev_extralarge = 0

            cumulative = reading['total']
            incremental = cumulative - prev_total

            incr_a3 = reading['a3'] - prev_a3
            incr_black = reading['black'] - prev_black
            incr_large = reading['large'] - prev_large
            incr_colour = reading['colour'] - prev_colour
            incr_extralarge = reading['extralarge'] - prev_extralarge

            # Calculate total from sum of sub-meters
            incr_total_from_subs = incr_a3 + incr_black + incr_large + incr_colour + incr_extralarge

            # Shift to next month to match Xerox reporting (they report on first day of next month)
            # Get next month for reporting
            if i + 1 < len(month_list):
                report_as_month = month_list[i + 1]
            else:
                report_as_month = None

            # Only store if the NEXT month is in our report range
            if incremental >= 0 and report_as_month and report_as_month in report_months:
                result[serial][report_as_month] = {
                    'volume': incr_total_from_subs,  # Use sum of sub-meters
                    'a3': incr_a3,
                    'black': incr_black,
                    'large': incr_large,
                    'colour': incr_colour,
                    'extralarge': incr_extralarge,
                    'model': reading['model'],
                    'company': reading['company'],
                    'cumulative': cumulative,
                    'meterreading_id': reading.get('meterreading_id')
                }

    return result


def query_all_databases() -> List[Tuple]:
    """Query all databases"""
    all_results = []
    print("Querying databases...")

    for i, company in enumerate(COMPANIES, 1):
        print(f"[{i}/{len(COMPANIES)}] {company['name']}...", end=" ")
        try:
            conn = mysql.connector.connect(
                host=company["host"],
                port=company["port"],
                user=DB_USER,
                password=DB_PASSWORD,
                database=company["schema"],
                connect_timeout=10,
                charset='utf8mb4',
                collation='utf8mb4_general_ci'  # Use older collation compatible with all MySQL versions
            )
            cursor = conn.cursor()
            cursor.execute(QUERY)
            results = cursor.fetchall()
            cursor.close()
            conn.close()

            for row in results:
                all_results.append((*row, company['name']))
            print(f"✓ {len(results)} rows")
        except Exception as e:
            print(f"✗ {str(e)[:50]}")

    return all_results


def create_excel_report(excel_data: Dict, db_data: Dict, raw_results: List[Tuple], device_meters: Dict, output_path: str):
    """Create Excel report - only serials in both systems"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Only serials in both systems
    common_serials = set(excel_data.keys()) & set(db_data.keys())

    # Single sheet: Comparison
    ws = wb.create_sheet("Volume Comparison", 0)

    # Styles for total columns
    total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    total_header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    total_header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Serial", "Model", "BMS Company", "Xerox Customer", "Month",
        "", # Spacer
        "BMS A3", "BMS Black", "BMS Large", "BMS Colour", "BMS XL", "BMS Total",
        "", # Spacer
        "Xerox A3 Mono", "Xerox A4 Mono", "Xerox A3 Color", "Xerox A4 Color", "Xerox Total",
        "", # Spacer
        "Difference", "BMS Balance"
    ]
    ws.append(headers)

    for idx, cell in enumerate(ws[1], 1):
        if cell.value in ["BMS Total", "Xerox Total"]:
            cell.fill = total_header_fill
            cell.font = total_header_font
        elif cell.value:
            cell.fill = header_fill
            cell.font = header_font

    # Month order for sorting
    month_order = {'Mar-2025': 1, 'Apr-2025': 2, 'May-2025': 3, 'Jun-2025': 4,
                   'Jul-2025': 5, 'Aug-2025': 6, 'Sep-2025': 7, 'Nov-2025': 8}

    for serial in sorted(common_serials):
        excel_months = excel_data[serial]
        db_months = db_data[serial]
        all_months = sorted(set(excel_months.keys()) | set(db_months.keys()),
                           key=lambda x: month_order.get(x, 99))

        # Get model and company from any available month for this serial
        serial_model = 'N/A'
        serial_company = 'N/A'
        for month_data in db_months.values():
            if month_data.get('model'):
                serial_model = month_data['model']
            if month_data.get('company'):
                serial_company = month_data['company']
            if serial_model != 'N/A' and serial_company != 'N/A':
                break

        # Get Xerox customer name from any available month for this serial
        xerox_customer = 'N/A'
        for month_data in excel_months.values():
            if month_data.get('customer'):
                xerox_customer = month_data['customer']
                break

        for month in all_months:
            db_info = db_months.get(month, {})
            xerox_info = excel_months.get(month, {})

            model = serial_model
            company = serial_company
            customer = xerox_customer

            bms_a3 = db_info.get('a3', 0)
            bms_black = db_info.get('black', 0)
            bms_large = db_info.get('large', 0)
            bms_colour = db_info.get('colour', 0)
            bms_xl = db_info.get('extralarge', 0)
            bms_total = db_info.get('volume', 0)
            bms_balance = db_info.get('cumulative', 0)

            xerox_a3_mono = xerox_info.get('a3_mono', 0)
            xerox_a4_mono = xerox_info.get('a4_mono', 0)
            xerox_a3_color = xerox_info.get('a3_color', 0)
            xerox_a4_color = xerox_info.get('a4_color', 0)
            xerox_total = xerox_info.get('total', 0)

            difference = bms_total - xerox_total

            row = [
                serial, model, company, customer, month,
                "", # Spacer
                bms_a3, bms_black, bms_large, bms_colour, bms_xl, bms_total,
                "", # Spacer
                xerox_a3_mono, xerox_a4_mono, xerox_a3_color, xerox_a4_color, xerox_total,
                "", # Spacer
                difference, bms_balance
            ]
            ws.append(row)

            # Highlight total columns
            row_num = ws.max_row
            ws[f'L{row_num}'].fill = total_fill  # BMS Total (now column L)
            ws[f'R{row_num}'].fill = total_fill  # Xerox Total (now column R)

    # ========== BMS Source Data Sheet (ALL DAILY READINGS) ==========
    ws_bms = wb.create_sheet("BMS Source Data", 1)

    bms_headers = [
        "Serial", "Model", "Company", "Reading Date", "Reading Time",
        "Meter Reading No", "Balance",
        "A3", "Black", "Large", "Colour", "XL"
    ]
    ws_bms.append(bms_headers)

    for cell in ws_bms[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add ALL daily readings from raw_results for common serials only
    # raw_results structure: (serial, model, status, date, datetime, total, a3, black, large, colour, xl, meterreadingid, meterreading_no, company)
    for row in raw_results:
        serial = row[0]

        # Only include serials that are in both systems
        if serial not in common_serials:
            continue

        model = row[1]
        reading_date = row[3]
        reading_datetime = row[4]
        total = row[5]
        a3 = row[6]
        black = row[7]
        large = row[8]
        colour = row[9]
        xl = row[10]
        meterreading_no = row[12]  # meterreading_no field
        company = row[13]

        bms_row = [
            serial, model, company, str(reading_date), str(reading_datetime),
            meterreading_no, total,
            a3, black, large, colour, xl
        ]
        ws_bms.append(bms_row)

    # ========== Xerox Source Data Sheet ==========
    ws_xerox = wb.create_sheet("Xerox Source Data", 2)

    xerox_headers = [
        "Serial", "Customer", "Month",
        "Xerox A3 Mono", "Xerox A4 Mono", "Xerox A3 Color", "Xerox A4 Color", "Xerox Total"
    ]
    ws_xerox.append(xerox_headers)

    for cell in ws_xerox[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add Xerox data for common serials
    for serial in sorted(common_serials):
        excel_months = excel_data[serial]

        # Get customer from any available month
        xerox_customer = 'N/A'
        for month_data in excel_months.values():
            if month_data.get('customer'):
                xerox_customer = month_data['customer']
                break

        all_months = sorted(excel_months.keys(), key=lambda x: month_order.get(x, 99))

        for month in all_months:
            xerox_info = excel_months[month]

            row = [
                serial, xerox_customer, month,
                xerox_info.get('a3_mono', 0),
                xerox_info.get('a4_mono', 0),
                xerox_info.get('a3_color', 0),
                xerox_info.get('a4_color', 0),
                xerox_info.get('total', 0)
            ]
            ws_xerox.append(row)

    # ========== Balance Reconciliation Sheet ==========
    # Simplified: Just compare TOTAL balances (Xerox Total vs BMS Balance field)
    # This handles both color printers (sub-meters) and B&W printers (Balance only)
    ws_recon = wb.create_sheet("Balance Recon", 3)

    recon_headers = [
        "Serial", "Xerox Company", "Xerox Product", "BMS Company", "BMS Model",
        "Xerox Date", "BMS Date",
        "",  # Spacer
        "Xerox Total", "BMS Balance", "Difference",
        "Calc Method", "Status"
    ]
    ws_recon.append(recon_headers)

    for cell in ws_recon[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Build BMS readings lookup by serial -> date -> readings
    # raw_results: (serial, model, status, date, datetime, total, a3, black, large, colour, xl, meterreadingid, meterreading_no, company)
    bms_by_serial_date = defaultdict(dict)
    for row in raw_results:
        serial = str(row[0]).strip() if row[0] else None
        if not serial:
            continue
        reading_date = row[3]
        if reading_date:
            date_key = reading_date.strftime("%Y-%m-%d") if hasattr(reading_date, 'strftime') else str(reading_date)[:10]
            # Store the reading - later ones overwrite earlier (we want latest reading per day)
            # Use the 'total' field from database (row[5]) - this is the Balance field
            bms_by_serial_date[serial][date_key] = {
                'model': row[1],
                'company': row[13],
                'balance': row[5] or 0,  # This is the Balance/total field from BMS
                'date': reading_date
            }

    # Process each serial in device_meters - ONLY include serials that exist in BMS
    for serial in sorted(device_meters.keys()):
        # Skip if serial not in BMS at all
        if serial not in bms_by_serial_date:
            continue

        xerox_info = device_meters[serial]
        xerox_date = xerox_info['last_date']

        if not xerox_date:
            continue

        xerox_date_str = xerox_date.strftime("%Y-%m-%d") if hasattr(xerox_date, 'strftime') else str(xerox_date)[:10]

        # Get the LATEST BMS reading (not matching Xerox date)
        bms_dates = bms_by_serial_date[serial]
        sorted_dates = sorted(bms_dates.keys())

        if not sorted_dates:
            continue

        # Use the latest BMS date
        bms_date_used = sorted_dates[-1]
        bms_info = bms_dates[bms_date_used]

        # Get Xerox total (sum of all sub-meters from Device Meters file)
        xerox_total = xerox_info.get('total', 0) or 0

        # Get BMS Balance (the total/balance field from database)
        bms_balance = bms_info.get('balance', 0) or 0
        bms_company = bms_info.get('company', 'N/A')
        bms_model = bms_info.get('model', 'N/A')

        # Get calc method
        calc_method = xerox_info.get('calc_method', 'Unknown')

        # Calculate difference
        difference = xerox_total - bms_balance

        # Determine status
        if calc_method == 'NO DATA':
            status = "NO XEROX DATA"
        elif difference == 0:
            status = "BALANCED"
        else:
            status = "VARIANCE"

        row = [
            serial, xerox_info.get('company', 'N/A'), xerox_info.get('product', 'N/A'),
            bms_company, bms_model,
            xerox_date_str, bms_date_used,
            "",  # Spacer
            xerox_total, bms_balance, difference,
            calc_method, status
        ]
        ws_recon.append(row)

        # Apply formatting
        row_num = ws_recon.max_row

        # Highlight difference column
        diff_cell = ws_recon.cell(row_num, 11)
        if diff_cell.value and diff_cell.value != 0:
            diff_cell.fill = red_fill
        elif diff_cell.value == 0:
            diff_cell.fill = green_fill

        # Highlight status
        status_cell = ws_recon.cell(row_num, 13)
        if status == "BALANCED":
            status_cell.fill = green_fill
        elif status == "VARIANCE":
            status_cell.fill = red_fill
        elif status == "NO XEROX DATA":
            status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange

    # Auto-size columns for all sheets
    for sheet in wb:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(output_path)

    # Print summary
    print(f"\n✓ Excel report saved: {output_path}")
    print(f"  Serials compared: {len(common_serials)}")
    total_rows = ws.max_row - 1  # Exclude header
    print(f"  Total comparisons: {total_rows}")


def create_balance_report(device_meters: Dict, raw_results: List[Tuple], output_path: str):
    """Create standalone Xerox vs BMS balance comparison report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Comparison"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    headers = [
        "Serial", "Xerox Company", "Xerox Product", "BMS Company", "BMS Model",
        "Xerox Date", "BMS Date",
        "",  # Spacer
        "Xerox Total", "BMS Balance", "Difference",
        "Calc Method", "Status"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Build BMS readings lookup by serial -> date -> readings
    bms_by_serial_date = defaultdict(dict)
    for row in raw_results:
        serial = str(row[0]).strip() if row[0] else None
        if not serial:
            continue
        reading_date = row[3]
        if reading_date:
            date_key = reading_date.strftime("%Y-%m-%d") if hasattr(reading_date, 'strftime') else str(reading_date)[:10]
            bms_by_serial_date[serial][date_key] = {
                'model': row[1],
                'company': row[13],
                'balance': row[5] or 0,
                'date': reading_date
            }

    # Process each serial in device_meters - ONLY include serials that exist in BMS
    for serial in sorted(device_meters.keys()):
        if serial not in bms_by_serial_date:
            continue

        xerox_info = device_meters[serial]
        xerox_date = xerox_info['last_date']

        if not xerox_date:
            continue

        xerox_date_str = xerox_date.strftime("%Y-%m-%d") if hasattr(xerox_date, 'strftime') else str(xerox_date)[:10]

        # Get the LATEST BMS reading
        bms_dates = bms_by_serial_date[serial]
        sorted_dates = sorted(bms_dates.keys())

        if not sorted_dates:
            continue

        bms_date_used = sorted_dates[-1]
        bms_info = bms_dates[bms_date_used]

        xerox_total = xerox_info.get('total', 0) or 0
        bms_balance = bms_info.get('balance', 0) or 0
        bms_company = bms_info.get('company', 'N/A')
        bms_model = bms_info.get('model', 'N/A')
        calc_method = xerox_info.get('calc_method', 'Unknown')

        difference = xerox_total - bms_balance

        if calc_method == 'NO DATA':
            status = "NO XEROX DATA"
        elif difference == 0:
            status = "BALANCED"
        else:
            status = "VARIANCE"

        row = [
            serial, xerox_info.get('company', 'N/A'), xerox_info.get('product', 'N/A'),
            bms_company, bms_model,
            xerox_date_str, bms_date_used,
            "",
            xerox_total, bms_balance, difference,
            calc_method, status
        ]
        ws.append(row)

        row_num = ws.max_row

        # Highlight difference column
        diff_cell = ws.cell(row_num, 11)
        if diff_cell.value and diff_cell.value != 0:
            diff_cell.fill = red_fill
        elif diff_cell.value == 0:
            diff_cell.fill = green_fill

        # Highlight status
        status_cell = ws.cell(row_num, 13)
        if status == "BALANCED":
            status_cell.fill = green_fill
        elif status == "VARIANCE":
            status_cell.fill = red_fill
        elif status == "NO XEROX DATA":
            status_cell.fill = orange_fill

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(output_path)
    print(f"✓ Balance report saved: {output_path}")
    print(f"  Serials compared: {ws.max_row - 1}")


def main():
    # Use paths relative to script location for portability
    script_dir = Path(__file__).parent
    # Check if Excel file exists in same directory as script (remote execution)
    excel_in_script_dir = script_dir / "Volumes from Xerox.xlsx"
    if excel_in_script_dir.exists():
        excel_path = excel_in_script_dir
        device_meters_path = script_dir / "Device Current Meters based on last Reading Date.xlsx"
        output_path = script_dir / "Volume_Comparison_Report.xlsx"
        balance_output_path = script_dir / "Xerox_vs_BMS_Balances.xlsx"
    else:
        # Local execution - files in project root
        project_root = script_dir.parent
        excel_path = project_root / "Volumes from Xerox.xlsx"
        device_meters_path = project_root / "Device Current Meters based on last Reading Date.xlsx"
        output_path = project_root / "Volume_Comparison_Report.xlsx"
        balance_output_path = project_root / "Xerox_vs_BMS_Balances.xlsx"

    print("Loading Excel data...")
    excel_data = load_excel_data(excel_path)
    print(f"✓ Loaded {len(excel_data)} serials from Excel\n")

    print("Loading Device Meters data...")
    device_meters = {}
    if device_meters_path.exists():
        device_meters = load_device_meters(device_meters_path)
        print(f"✓ Loaded {len(device_meters)} serials from Device Meters\n")
    else:
        print("⚠ Device Meters file not found - skipping balance recon\n")

    raw_results = query_all_databases()
    print(f"\n✓ Retrieved {len(raw_results)} total rows\n")

    print("Calculating incremental volumes...")
    db_data = calculate_incremental_volumes(raw_results)
    print(f"✓ Processed {len(db_data)} unique serials\n")

    print("Creating Excel report...")
    create_excel_report(excel_data, db_data, raw_results, device_meters, output_path)

    # Create separate balance report
    if device_meters:
        print("\nCreating balance comparison report...")
        create_balance_report(device_meters, raw_results, balance_output_path)


if __name__ == "__main__":
    main()
