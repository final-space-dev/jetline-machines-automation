#!/usr/bin/env python3
"""
Detailed analysis of serial 3135455511 showing raw data and calculations
"""

import mysql.connector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import OrderedDict, defaultdict
import os
from dotenv import load_dotenv

load_dotenv()

DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

if not DB_USER or not DB_PASSWORD:
    raise ValueError("DB_USER and DB_PASSWORD must be set in .env file")

def analyze_serial():
    # Query Fixtrade for serial 3135455511
    conn = mysql.connector.connect(
        host='172.20.251.127',
        port=3306,
        user=DB_USER,
        password=DB_PASSWORD,
        database='nscbms2'
    )

    cursor = conn.cursor()
    cursor.execute('''
    SELECT
        mrd.reading_date,
        crme.createdtime as reading_date_time,
        mrd.total
    FROM bms_machines ma
    LEFT JOIN vtiger_crmentity crm on crm.crmid = ma.machinesid
    LEFT JOIN bms_meterreading mrd on mrd.asset = ma.machinesid
    LEFT JOIN vtiger_crmentity crme on crme.crmid = mrd.meterreadingid
    WHERE ma.serialnumber = "3135455511"
        AND crm.deleted <> 1
        AND ma.machinestatus = 1
        AND YEAR(mrd.reading_date) = 2025
        AND MONTH(mrd.reading_date) IN (3,4,5,6,7,8,9,10,11)
    ORDER BY mrd.reading_date, crme.createdtime
    ''')

    raw_data = cursor.fetchall()
    cursor.close()
    conn.close()

    # Process data using Qlik logic
    by_date = defaultdict(list)
    for read_date, read_time, total in raw_data:
        date_key = read_date.strftime("%Y-%m-%d")
        by_date[date_key].append((read_time, total))

    # Get last reading per day
    daily_last = {}
    daily_max = {}
    for date_key, readings in by_date.items():
        readings.sort(key=lambda x: x[0] if x[0] else datetime.min, reverse=True)
        daily_last[date_key] = readings[0][1]  # Last by time
        daily_max[date_key] = max(r[1] for r in readings)  # Max value

    # Group by month
    monthly_last = OrderedDict()
    monthly_max = OrderedDict()

    for date_str in sorted(daily_last.keys()):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_year = date_obj.strftime("%b-%Y")
        monthly_last[month_year] = daily_last[date_str]
        if month_year not in monthly_max or daily_max[date_str] > monthly_max[month_year]:
            monthly_max[month_year] = daily_max[date_str]

    # Load Excel data
    wb_excel = openpyxl.load_workbook('/Users/alwynkotze/Documents/JDW/jetline-machines-automation/Volumes from Xerox.xlsx')
    ws_excel = wb_excel.active
    xerox_data = {}
    for row in ws_excel.iter_rows(min_row=3, values_only=True):
        customer, model, serial, month_year, volume = row
        if serial == '3135455511' and month_year and month_year != 'Total':
            xerox_data[month_year] = volume

    # Create analysis Excel
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary Analysis"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Summary sheet
    ws_summary.append(["Analysis: Serial 3135455511 (Xerox Iridesse - Fixtrade)"])
    ws_summary.append([])
    ws_summary['A1'].font = Font(bold=True, size=14)

    headers = ["Month", "BMS End Balance (LAST)", "BMS End Balance (MAX)", "BMS Incremental (LAST)",
               "BMS Incremental (MAX)", "Xerox Volume", "Diff vs LAST", "Diff vs MAX", "Status"]
    ws_summary.append(headers)

    for cell in ws_summary[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    prev_last = 0
    prev_max = 0

    months = ['Mar-2025', 'Apr-2025', 'May-2025', 'Jun-2025', 'Jul-2025', 'Aug-2025', 'Sep-2025', 'Oct-2025', 'Nov-2025']

    for month in months:
        if month not in monthly_last:
            continue

        end_last = monthly_last[month]
        end_max = monthly_max.get(month, end_last)

        incr_last = end_last - prev_last
        incr_max = end_max - prev_max

        xerox_vol = xerox_data.get(month, 0)

        diff_last = incr_last - xerox_vol
        diff_max = incr_max - xerox_vol

        # Determine which method matches better
        if abs(diff_last) < abs(diff_max):
            status = "LAST method better"
            fill = green_fill if abs(diff_last) < 1000 else yellow_fill
        else:
            status = "MAX method better"
            fill = red_fill if abs(diff_max) > 10000 else yellow_fill

        row = [month, end_last, end_max, incr_last, incr_max, xerox_vol, diff_last, diff_max, status]
        ws_summary.append(row)

        last_row = ws_summary.max_row
        ws_summary[f'I{last_row}'].fill = fill

        prev_last = end_last
        prev_max = end_max

    # Raw data sheet
    ws_raw = wb.create_sheet("Raw Daily Readings")
    ws_raw.append(["Date", "Time", "Cumulative Total", "Month", "Is Last of Day", "Is Max of Day"])

    for cell in ws_raw[1]:
        cell.fill = header_fill
        cell.font = header_font

    for read_date, read_time, total in raw_data:
        date_key = read_date.strftime("%Y-%m-%d")
        month = read_date.strftime("%b-%Y")
        is_last = (daily_last.get(date_key) == total)
        is_max = (daily_max.get(date_key) == total)

        ws_raw.append([
            read_date.strftime("%Y-%m-%d"),
            read_time.strftime("%H:%M:%S") if read_time else "N/A",
            total,
            month,
            "YES" if is_last else "",
            "YES" if is_max else ""
        ])

        if is_last:
            ws_raw[f'E{ws_raw.max_row}'].fill = green_fill
        if is_max:
            ws_raw[f'F{ws_raw.max_row}'].fill = yellow_fill

    # Auto-size columns
    for ws in wb:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output_path = "/Users/alwynkotze/Documents/JDW/jetline-machines-automation/Serial_3135455511_Analysis.xlsx"
    wb.save(output_path)
    print(f"✓ Analysis saved: {output_path}")

if __name__ == "__main__":
    analyze_serial()
