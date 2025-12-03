#!/usr/bin/env python3
"""
Analysis including all meter subtotals (a3, black, colour, large, extralarge)
to see if Xerox is calculating differently
"""

import mysql.connector
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import OrderedDict, defaultdict
import os
from dotenv import load_dotenv

load_dotenv()

DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

if not DB_USER or not DB_PASSWORD:
    raise ValueError("DB_USER and DB_PASSWORD must be set in .env file")

def analyze_with_subtotals():
    conn = mysql.connector.connect(
        host='172.20.251.127',
        port=3306,
        user=DB_USER,
        password=DB_PASSWORD,
        database='nscbms2'
    )

    cursor = conn.cursor()
    cursor.execute('''
    SELECT
        mrd.reading_date,
        crme.createdtime,
        mrd.total,
        mrd.a3,
        mrd.black,
        mrd.large,
        mrd.colour,
        mrd.extralarge
    FROM bms_machines ma
    LEFT JOIN vtiger_crmentity crm on crm.crmid = ma.machinesid
    LEFT JOIN bms_meterreading mrd on mrd.asset = ma.machinesid
    LEFT JOIN vtiger_crmentity crme on crme.crmid = mrd.meterreadingid
    WHERE ma.serialnumber = "3135455511"
        AND crm.deleted <> 1
        AND ma.machinestatus = 1
        AND YEAR(mrd.reading_date) = 2025
        AND MONTH(mrd.reading_date) IN (3,4,5,6,7,8)
    ORDER BY mrd.reading_date, crme.createdtime
    ''')

    raw_data = cursor.fetchall()
    cursor.close()
    conn.close()

    # Process by date - take last reading per day
    by_date = defaultdict(list)
    for read_date, read_time, total, a3, black, large, colour, extralarge in raw_data:
        date_key = read_date.strftime("%Y-%m-%d")

        # Calculate sum of sub-meters (if they exist)
        sub_sum = 0
        for val in [a3, black, large, colour, extralarge]:
            if val is not None:
                sub_sum += val

        by_date[date_key].append({
            'time': read_time,
            'total': total or 0,
            'a3': a3 or 0,
            'black': black or 0,
            'large': large or 0,
            'colour': colour or 0,
            'extralarge': extralarge or 0,
            'sub_sum': sub_sum
        })

    # Get last reading per day
    daily_last = {}
    for date_key, readings in by_date.items():
        readings.sort(key=lambda x: x['time'] if x['time'] else datetime.min, reverse=True)
        daily_last[date_key] = readings[0]

    # Group by month
    monthly_last = OrderedDict()
    for date_str in sorted(daily_last.keys()):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_year = date_obj.strftime("%b-%Y")
        monthly_last[month_year] = daily_last[date_str]

    # Load Excel data
    wb_excel = openpyxl.load_workbook('/Users/alwynkotze/Documents/JDW/jetline-machines-automation/Volumes from Xerox.xlsx')
    ws_excel = wb_excel.active
    xerox_data = {}
    for row in ws_excel.iter_rows(min_row=3, values_only=True):
        customer, model, serial, month_year, volume = row
        if serial == '3135455511' and month_year and month_year != 'Total':
            xerox_data[month_year] = volume

    # Create Excel report
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meter Analysis"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(["Serial 3135455511 - Detailed Meter Analysis"])
    ws.append([])
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Month", "DB Balance", "DB Movement", "Xerox Movement", "Difference"]
    ws.append(headers)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font

    prev_total = 0

    months = ['Mar-2025', 'Apr-2025', 'May-2025', 'Jun-2025', 'Jul-2025', 'Aug-2025']

    for month in months:
        if month not in monthly_last:
            continue

        data = monthly_last[month]
        db_balance = data['total']
        db_movement = db_balance - prev_total
        xerox_movement = xerox_data.get(month, 0)
        difference = db_movement - xerox_movement

        row = [
            month,
            db_balance,
            db_movement,
            xerox_movement,
            difference
        ]
        ws.append(row)

        prev_total = db_balance

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output_path = "/Users/alwynkotze/Documents/JDW/jetline-machines-automation/Meter_Subtotals_Analysis.xlsx"
    wb.save(output_path)
    print(f"✓ Analysis saved: {output_path}")

if __name__ == "__main__":
    analyze_with_subtotals()
