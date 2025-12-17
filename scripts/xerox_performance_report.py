#!/usr/bin/env python3
"""
Xerox Performance Report - Combines volumes, balance, utilization, and service calls
Xerox data + BMS machine info (category, install_date, company)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict
import json


def load_volumes(filepath: str) -> dict:
    """Load monthly volumes by serial number"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Structure: {serial: {customer, product, months: {month: total}}}
    data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer, model, serial, month_year, a3_mono, a4_mono, a3_color, a4_color, total = row
        if not serial:
            continue
        serial = str(serial).strip()

        if serial not in data:
            data[serial] = {
                'customer': customer,
                'product': model,
                'months': {}
            }

        if month_year:
            data[serial]['months'][month_year] = total or 0

    return data


def load_device_meters(filepath: str) -> dict:
    """Load balance by serial number (using same logic as export script)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    data = {}

    meter_map = {
        'Black Impressions': 'black',
        'Color Impressions': 'color',
        'Total Impressions': 'total_impressions'
    }

    for row_num in range(2, ws.max_row + 1):
        company = ws.cell(row_num, 1).value
        product = ws.cell(row_num, 2).value
        serial = str(ws.cell(row_num, 3).value).strip() if ws.cell(row_num, 3).value else None
        meter_desc = ws.cell(row_num, 4).value
        reading = ws.cell(row_num, 5).value
        last_date = ws.cell(row_num, 6).value

        if not serial or meter_desc not in meter_map:
            continue

        if serial not in data:
            data[serial] = {
                'company': company,
                'product': product,
                'last_date': last_date,
                'black': 0,
                'color': 0,
                'total_impressions': 0
            }

        field = meter_map[meter_desc]
        data[serial][field] = reading or 0

        if last_date and (not data[serial]['last_date'] or last_date > data[serial]['last_date']):
            data[serial]['last_date'] = last_date

    # Calculate total
    for serial in data:
        d = data[serial]
        total_impressions = d.get('total_impressions', 0) or 0
        black = d.get('black', 0) or 0
        color = d.get('color', 0) or 0

        if total_impressions > 0:
            d['balance'] = total_impressions
        else:
            d['balance'] = black + color

    return data


def load_mttr(color_path: str, mono_path: str) -> dict:
    """Load service calls by Product + Customer (combine color and mono)"""
    data = defaultdict(lambda: {'calls': 0})

    # Load color MTTR
    wb = openpyxl.load_workbook(color_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        product, customer, avg_mttr, calls = row
        if customer == 'Total' or not customer:
            continue
        key = (product, customer)
        data[key]['calls'] += calls or 0

    # Load mono MTTR
    wb = openpyxl.load_workbook(mono_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        product, customer, avg_mttr, calls = row
        if customer == 'Total' or not customer:
            continue
        key = (product, customer)
        data[key]['calls'] += calls or 0

    return dict(data)


def load_utilization(color_path: str, mono_path: str) -> dict:
    """Load utilization by Product + Customer (take max of color/mono)"""
    data = {}

    # Load color utilization
    wb = openpyxl.load_workbook(color_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        product, customer, utilization = row
        if customer == 'Total' or not customer:
            continue
        key = (product, customer)
        data[key] = utilization or 0

    # Load mono utilization (take max if already exists)
    wb = openpyxl.load_workbook(mono_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        product, customer, utilization = row
        if customer == 'Total' or not customer:
            continue
        key = (product, customer)
        if key in data:
            data[key] = max(data[key], utilization or 0)
        else:
            data[key] = utilization or 0

    return data


def load_bms_machine_info(filepath: str) -> dict:
    """Load BMS machine info (category, install_date, company) from JSON"""
    if not Path(filepath).exists():
        return {}
    with open(filepath, 'r') as f:
        return json.load(f)


def create_performance_report(volumes: dict, balances: dict, mttr: dict, utilization: dict, bms_info: dict, output_path: str):
    """Create the Xerox Performance Report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    number_format = '#,##0'
    pct_format = '0.0%'

    # All months (financial YTD: Mar-Nov 2025)
    target_months = ['Mar-2025', 'Apr-2025', 'May-2025', 'Jun-2025', 'Jul-2025', 'Aug-2025', 'Sep-2025', 'Oct-2025', 'Nov-2025']

    # Headers - with BMS columns between Customer Name and Product
    headers = [
        'Customer Name (Xerox)', 'BMS Company', 'Category', 'Install Date', 'Product', 'Serial',
        'Mar-2025', 'Apr-2025', 'May-2025', 'Jun-2025', 'Jul-2025', 'Aug-2025', 'Sep-2025', 'Oct-2025', 'Nov-2025',
        'YTD Volume', 'Balance', 'Utilization %', 'Service Calls', 'Volume/Call'
    ]
    ws.append(headers)

    for idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Process each serial
    for serial in sorted(volumes.keys()):
        vol_data = volumes[serial]
        customer = vol_data['customer']
        product = vol_data['product']
        months = vol_data['months']

        # Get BMS machine info
        machine = bms_info.get(serial, {})
        bms_company = machine.get('bms_company', '')
        category = machine.get('category', '')
        install_date = machine.get('install_date', '')

        # Get monthly volumes
        month_volumes = [months.get(m, 0) or 0 for m in target_months]
        ytd_volume = sum(month_volumes)

        # Get balance from device meters
        balance = 0
        if serial in balances:
            balance = balances[serial].get('balance', 0) or 0

        # Get utilization and calls by product + customer
        key = (product, customer)
        util_pct = utilization.get(key, 0) or 0
        calls = mttr.get(key, {}).get('calls', 0) or 0

        # Calculate volume per call
        vol_per_call = ytd_volume / calls if calls > 0 else 0

        row = [
            customer, bms_company, category, install_date, product, serial,
            month_volumes[0], month_volumes[1], month_volumes[2],
            month_volumes[3], month_volumes[4], month_volumes[5],
            month_volumes[6], month_volumes[7], month_volumes[8],
            ytd_volume, balance, util_pct, calls, vol_per_call
        ]
        ws.append(row)

        row_num = ws.max_row

        # Format numbers (columns 7-17: all month volumes, YTD, balance)
        for col in range(7, 18):
            ws.cell(row_num, col).number_format = number_format

        # Format utilization as percentage (column 18)
        ws.cell(row_num, 18).number_format = pct_format

        # Format calls (column 19)
        ws.cell(row_num, 19).number_format = number_format

        # Format volume/call (column 20)
        ws.cell(row_num, 20).number_format = number_format

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"✓ Performance report saved: {output_path}")
    print(f"  Machines: {ws.max_row - 1}")


def main():
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    # Input files
    volumes_path = project_root / "Volumes from Xerox.xlsx"
    device_meters_path = project_root / "Device Current Meters based on last Reading Date.xlsx"
    color_mttr_path = project_root / "Color MTTR by Product and Site.xlsx"
    mono_mttr_path = project_root / "Mono MTTR by Product and Site.xlsx"
    color_util_path = project_root / "Color Utilization by Product and Site.xlsx"
    mono_util_path = project_root / "Mono Utilization by Product and Site.xlsx"
    bms_info_path = project_root / "bms_machine_info.json"

    # Output file
    output_path = project_root / "Xerox_Performance_Report.xlsx"

    print("Loading Xerox data...")

    print("  Loading volumes...")
    volumes = load_volumes(volumes_path)
    print(f"  ✓ {len(volumes)} serials")

    print("  Loading device meters (balances)...")
    balances = load_device_meters(device_meters_path)
    print(f"  ✓ {len(balances)} serials")

    print("  Loading MTTR (service calls)...")
    mttr = load_mttr(color_mttr_path, mono_mttr_path)
    print(f"  ✓ {len(mttr)} product/customer combinations")

    print("  Loading utilization...")
    utilization = load_utilization(color_util_path, mono_util_path)
    print(f"  ✓ {len(utilization)} product/customer combinations")

    print("  Loading BMS machine info...")
    bms_info = load_bms_machine_info(bms_info_path)
    print(f"  ✓ {len(bms_info)} serials")

    print("\nCreating performance report...")
    create_performance_report(volumes, balances, mttr, utilization, bms_info, output_path)


if __name__ == "__main__":
    main()
